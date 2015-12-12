import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
import glob

final_file = Workbook()
ws = final_file.active
nextrow = 1
files = glob.glob("C:\\Users\\Jarrod\PycharmProjects\\untitled\\empty*.xlsx")

for i in files:
    if i.endswith(".xlsx"):
        wb = openpyxl.load_workbook(i)
        sheet = wb.get_active_sheet()
        columns = sheet.get_highest_column
        if nextrow == 1:
            for row in range(1, sheet.get_highest_row()+1):
                for col in range(1, sheet.get_highest_column()+1):
                    colletter = get_column_letter(col)
                    ws[colletter + str(row)] = sheet[colletter + str(row)].value
                nextrow = row

        else:
            for row in range(2, sheet.get_highest_row()+1):
                for col in range(1, sheet.get_highest_column()+1):
                    colletter = get_column_letter(col)
                    ws[colletter + str(nextrow)] = sheet[colletter + str(row)].value
                nextrow = nextrow + 1
    else:
        continue
final_file.save("totaltrades.xlsx")
